import os
import pandas as pd
from openpyxl import load_workbook

# Ruta donde están los archivos insumo
RUTA_DATOS = "./insumos"
# Ruta de la plantilla base
PLANTILLA_BASE = "plantilla.xlsx"
# Carpeta de salida
SALIDA = "./salidas"

# Mapeo de columnas a celdas en la plantilla
MAPEO_CELDAS = {
    "A": "L11",
    "B": "C26",
    "C": "D26",
    "D": "F26",
    "E": "G33",
    "F": "H50",
    "G": "F21",
    "H": "L21"
}

# Iteramos por todos los archivos en la carpeta de insumos
for archivo in os.listdir(RUTA_DATOS):
    if archivo.endswith(".xlsx"):
        ruta_archivo = os.path.join(RUTA_DATOS, archivo)
        print(f"Procesando archivo: {archivo}")

        # Leer los datos ignorando la primera fila
        df = pd.read_excel(ruta_archivo, skiprows=1, header=None)

        # Crear carpeta de salida si no existe
        carpeta_salida = os.path.join(SALIDA, os.path.splitext(archivo)[0])
        os.makedirs(carpeta_salida, exist_ok=True)

        for i, fila in df.iterrows():
            # Cargar plantilla
            wb = load_workbook(PLANTILLA_BASE)
            ws = wb.active

            # Llenar celdas con datos
            for idx, col in enumerate("ABCDEFGH"):
                valor = fila[idx]
                celda = MAPEO_CELDAS[col]
                ws[celda] = valor

            # Construir nombre del archivo de salida
            nombre_archivo = f"ESAR__{fila[0]}__{fila[1]}__{fila[2]}.xlsx"
            ruta_salida = os.path.join(carpeta_salida, nombre_archivo)

            # Guardar archivo
            wb.save(ruta_salida)

print("✅ Todos los archivos han sido generados.")