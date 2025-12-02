import os
import io
from openpyxl import load_workbook
from PIL import Image

# ————————————————————————————————————————————————————————————————
# 1) RUTA EXACTA DE TU ARCHIVO Excel
ruta_excel = r"D:\MSI\INVENTARIOS RESUMEN\CLARO\GUIAS ENTREGAS 2024 PARA SUSTENTO 2023\Excel Guia\Formato de Devoluciones HUAWEI-MSI-Claro - EQUIPAMIENTO COMPLETO (1).xlsx"

# 2) CARPETA DONDE SE GUARDARÁN LAS IMÁGENES
carpeta_destino = r"D:\MSI\INVENTARIOS RESUMEN\CLARO\GUIAS ENTREGAS 2024 PARA SUSTENTO 2023\FOTOS DE CLARO"
# ————————————————————————————————————————————————————————————————

# Abrimos el libro en modo solo lectura de valores
wb = load_workbook(ruta_excel, data_only=True)

# Nombres de las pestañas (asegúrate de que coincidan exactamente)
hoja_formato = wb["FORMATO"]
hoja_fotos   = wb["Fotos de los Equipos"]

# Creamos la carpeta destino si no existe
os.makedirs(carpeta_destino, exist_ok=True)

def limpiar_nombre(texto: str) -> str:
    """
    Elimina caracteres que no pueden ir en un nombre de archivo en Windows.
    """
    invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
    for ch in invalid_chars:
        texto = texto.replace(ch, '')
    return texto.strip()

# ————————————————————————————————————————————————————————————————
# 3) En tu caso, ya sabemos que:
#    - Ítem 1 está en la fila 20, columna B (es decir, fila=20, columna=2).
#    - Ítem 125 está en la fila 144, columna B (20 + 124 = 144).
#
#    Por tanto, configuramos directamente fila_item1 = 20.
fila_item1 = 20
print(f"► Usando fila fija: Ítem 1 en fila {fila_item1}, Ítem 125 en fila {fila_item1 + 124}.")


# ————————————————————————————————————————————————————————————————
# 4) CARGAMOS LAS IMÁGENES de “Fotos de los Equipos” en un diccionario
#    La clave será (fila, columna) en la hoja, y el valor es lista de img
imagenes = {}
for img in hoja_fotos._images:
    fila_img = img.anchor._from.row + 1  # conv. 0-based → 1-based
    col_img  = img.anchor._from.col + 1
    coord = (fila_img, col_img)
    imagenes.setdefault(coord, []).append(img)


# ————————————————————————————————————————————————————————————————
# 5) RECORREMOS los ítems del 1 al 125:
#    - Para cada Ítem N, la fila en “FORMATO” será: fila_item1 + (N-1)
#    - En esa fila:
#         · CÓDIGO SAP  = columna C (índice 3)
#         · TEXTO BREVE = columna D (índice 4)
#    - En la hoja “Fotos de los Equipos” buscamos las imágenes en la MISMA fila
#      y en columnas B=2, C=3, D=4. 

for item in range(1, 126):
    fila_formato_item = fila_item1 + (item - 1)

    # Leemos CÓDIGO SAP desde columna C (3) y DESCRIPCIÓN desde columna D (4)
    valor_codigo_sap  = hoja_formato.cell(row=fila_formato_item, column=3).value
    valor_descripcion = hoja_formato.cell(row=fila_formato_item, column=4).value

    if not valor_codigo_sap or not valor_descripcion:
        print(f"Ítem {item:03d}: falta CÓDIGO SAP o DESCRIPCIÓN → se omite")
        continue

    codigo_sap  = limpiar_nombre(str(valor_codigo_sap))
    descripcion = limpiar_nombre(str(valor_descripcion))

    # En “Fotos de los Equipos”, asumimos que las 3 fotos del Ítem N están en:
    #  fila = fila_formato_item
    #  columnas B=2, C=3, D=4
    fila_fotos_item = fila_formato_item

    for indice_col, columna_foto in enumerate([2, 3, 4], start=1):
        lista_imgs = imagenes.get((fila_fotos_item, columna_foto), [])

        # Si hay más de una imagen en la misma celda, las numeramos con _1, _2, …
        for contador, img in enumerate(lista_imgs, start=1):
            # Nombre base: "<DESCRIPCIÓN> - <CÓDIGO_SAP> .<númeroFoto>"
            # Si hay >1 en la misma celda, añado "_<contador>" antes de .png
            if len(lista_imgs) > 1:
                nombre_archivo = f"{descripcion} - {codigo_sap} .{indice_col}_{contador}.png"
            else:
                nombre_archivo = f"{descripcion} - {codigo_sap} .{indice_col}.png"

            ruta_guardado = os.path.join(carpeta_destino, nombre_archivo)

            blob = img._data()
            if isinstance(blob, bytes):
                # Convertimos bytes a PIL.Image
                imagen_pil = Image.open(io.BytesIO(blob))
            else:
                # Ya viene como PIL.Image
                imagen_pil = blob

            # Guardamos la imagen como PNG
            imagen_pil.save(ruta_guardado)
            print(f"Ítem {item:03d} → guardada foto: {nombre_archivo}")

print("\n► ¡Extracción completada! Verifica la carpeta:")
print(f"   {carpeta_destino}")
