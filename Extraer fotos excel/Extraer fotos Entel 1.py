import os
import io
from openpyxl import load_workbook
from PIL import Image

# ————————————————————————————————————————————————————————————————
# 1) RUTA EXACTA DE TU ARCHIVO EXCEL
ruta_excel = r"D:\MSI\MBB ADICIONALES SUSTENTOS\Evidencias ordenado\0130180_LM_Las_Vegas\Fotos extraccion"

# 2) CARPETA DONDE GUARDARÁS LAS IMÁGENES EXTRAÍDAS
carpeta_destino = r"D:\MSI\INVENTARIOS EXCEL\Entel FOTOS nombradas"
# ————————————————————————————————————————————————————————————————

# Abrimos el Excel (solo lectura de valores)
wb = load_workbook(ruta_excel, data_only=True)

# ---------------------------------------------------------------------------------------------------
# 3) ELIGE la hoja que contiene las fotos y los datos. 
#    Si es la primera (activa) o tiene un nombre específico, cámbialo aquí.
#
#    Ejemplo:
#    hoja_datos = wb["NombreDeLaHoja"]
#
#    Si no estás seguro cuál es el nombre, puedes descomentar la línea de abajo 
#    para ver en consola el listado de hojas: 
#
#    print("Hojas disponibles en el libro:", wb.sheetnames)
#
#    Y luego sustituir "NombreDeLaHoja" por la que corresponda.
# ---------------------------------------------------------------------------------------------------
hoja_datos = wb.active
# ---------------------------------------------------------------------------------------------------

# Creamos la carpeta destino si no existe
os.makedirs(carpeta_destino, exist_ok=True)

def limpiar_nombre(texto: str) -> str:
    """
    Elimina caracteres inválidos para nombres de archivo (Windows).
    """
    invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
    for ch in invalid_chars:
        texto = texto.replace(ch, '')
    return texto.strip()

# ————————————————————————————————————————————————————————————————
# 4) CARGAMOS TODAS las imágenes de la hoja en un diccionario:
#    La clave será (fila, columna) y el valor una lista con los objetos image.
#    openpyxl guarda cada imagen en hoja_datos._images, con
#    img.anchor._from.row y img.anchor._from.col (cero-based).
#    Sumamos +1 para convertirlo a Excel (1-based).
imagenes = {}
for img in hoja_datos._images:
    fila_img = img.anchor._from.row + 1
    col_img  = img.anchor._from.col + 1
    coord = (fila_img, col_img)
    imagenes.setdefault(coord, []).append(img)

# ————————————————————————————————————————————————————————————————
# 5) RECORREMOS LAS FILAS 3 A 123 (ítems 1 a 121 reales, pero si hay solo 125 filas,
#    adaptamos al “123” que pediste). Ajusta el rango si fuera necesario.
#
#    Para cada fila (3..123):
#      - Columna O  = índice 15  → DESCRIPCIÓN
#      - Columna P  = índice 16  → FAMILIA (fallback)
#      - Columna R  = índice 18  → SERIE
#
#    Y las fotos en:
#      - Columna AB = índice 28  (número_foto = 1)
#      - Columna AC = índice 29  (número_foto = 2)
#      - Columna AD = índice 30  (número_foto = 3)
#
#    (Ten en cuenta que A=1, B=2, …, Z=26, AA=27, AB=28, AC=29, AD=30).
# ————————————————————————————————————————————————————————————————
for fila in range(3, 124):  # 3..123 inclusive
    # Leemos valores de DESCRIPCIÓN, FAMILIA y SERIE
    val_descripcion = hoja_datos.cell(row=fila, column=15).value  # columna O
    val_familia     = hoja_datos.cell(row=fila, column=16).value  # columna P
    val_serie       = hoja_datos.cell(row=fila, column=18).value  # columna R

    # Determinamos el texto base para el nombre:
    # Si existen descripción (O) y serie (R), usamos "<O> - <R>"
    # Si falta alguno, usamos sólo la familia (P)
    if val_descripcion and val_serie:
        descripcion = limpiar_nombre(str(val_descripcion))
        serie       = limpiar_nombre(str(val_serie))
        prefijo = f"{descripcion} - {serie}"
    else:
        # Si falta DESCRIPCIÓN o falta SERIE, usamos FAMILIA:
        if val_familia:
            prefijo = limpiar_nombre(str(val_familia))
        else:
            # Si no hay familia, ponemos genérico con la fila para que no colapse:
            prefijo = f"ItemFila{fila}"

    # Ahora, por cada una de las 3 columnas AB, AC, AD, guardamos la(s) foto(s):
    for indice_col, col_num in enumerate([28, 29, 30], start=1):
        fotos_en_celda = imagenes.get((fila, col_num), [])

        # Si la lista trae varias imágenes en la celda, las enumeramos con _1, _2…
        for contador, img in enumerate(fotos_en_celda, start=1):
            if len(fotos_en_celda) > 1:
                # Ejemplo: "<prefijo> .<numeroFoto>_<contador>.png"
                nombre_archivo = f"{prefijo} .{indice_col}_{contador}.png"
            else:
                # Ejemplo: "<prefijo> .<numeroFoto>.png"
                nombre_archivo = f"{prefijo} .{indice_col}.png"

            ruta_guardar = os.path.join(carpeta_destino, nombre_archivo)

            # Extraemos los bytes o PIL.Image de la imagen original
            blob = img._data()
            if isinstance(blob, bytes):
                imagen_pil = Image.open(io.BytesIO(blob))
            else:
                imagen_pil = blob  # ya es PIL.Image

            # Guardamos como PNG
            imagen_pil.save(ruta_guardar)
            print(f"Fila {fila} → guardada foto: {nombre_archivo}")

print("\n► ¡Proceso completado! Revisa la carpeta:")
print(f"   {carpeta_destino}")
